"""Proposal and document generation API endpoints."""

from __future__ import annotations

import io
from uuid import UUID

import structlog
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse

from app.core.http import content_disposition
from app.core.proposal_generator import ProposalGenerator
from app.dependencies import get_current_user, get_supabase_client
from app.middleware.rbac import require_permission
from app.models.pricing import IntegratedPricingInput, IntegratedPricingResult

logger = structlog.get_logger()
router = APIRouter(prefix="/api/v1/proposals", tags=["proposals"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _extract_vehicle_info(input_data: dict) -> dict:
    """Extract vehicle info dict from simulation input data."""
    return {
        "maker": input_data.get("maker", ""),
        "model": input_data.get("model", ""),
        "registration_year_month": input_data.get("registration_year_month", ""),
        "mileage_km": input_data.get("mileage_km", 0),
        "vehicle_class": input_data.get("vehicle_class", ""),
        "body_type": input_data.get("body_type", ""),
        "lease_term_months": input_data.get("lease_term_months", 36),
    }


def _load_simulation(supabase, simulation_id: UUID) -> dict:
    """Load a simulation row from Supabase; raise 404 if missing."""
    result = (
        supabase.table("simulations")
        .select("*")
        .eq("id", str(simulation_id))
        .single()
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=404, detail="Simulation not found")
    return result.data


# ---------------------------------------------------------------------------
# POST /generate – produce proposal PDF or HTML preview
# ---------------------------------------------------------------------------


@router.post("/generate")
async def generate_proposal(
    request: Request,
    simulation_id: UUID = None,
    user=Depends(require_permission("simulations", "read")),
    supabase=Depends(get_supabase_client),
):
    """Generate proposal PDF from simulation results.

    Returns PDF by default.  When the ``Accept`` header contains
    ``text/html`` or the ``HX-Request`` header is present the endpoint
    returns an HTML preview instead.
    """
    if not simulation_id:
        raise HTTPException(status_code=400, detail="simulation_id required")

    sim = _load_simulation(supabase, simulation_id)

    fund_info = {"fund_name": "カーチスファンド"}

    input_data = sim.get("input_data", {})
    result_data = sim.get("result", {})

    generator = ProposalGenerator()
    vehicle_info = _extract_vehicle_info(input_data)

    # Check Accept header – return HTML when requested by browser / HTMX
    accept = request.headers.get("accept", "")
    if "text/html" in accept or request.headers.get("HX-Request") == "true":
        html = generator.generate_html_preview(result_data, vehicle_info, fund_info)
        return HTMLResponse(content=html)

    pdf_bytes = generator.generate_pdf(result_data, vehicle_info, fund_info)

    # Detect whether we got a real PDF or an HTML fallback
    is_pdf = pdf_bytes[:5] == b"%PDF-"
    media = "application/pdf" if is_pdf else "text/html; charset=utf-8"
    ext = "pdf" if is_pdf else "html"

    filename = (
        f"proposal_{input_data.get('maker', 'vehicle')}"
        f"_{input_data.get('model', '')}.{ext}"
    )

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type=media,
        headers={"Content-Disposition": content_disposition(filename)},
    )


# ---------------------------------------------------------------------------
# GET /preview/{simulation_id} – HTML preview
# ---------------------------------------------------------------------------


@router.get("/preview/{simulation_id}")
async def preview_proposal(
    simulation_id: UUID,
    request: Request,
    user=Depends(require_permission("simulations", "read")),
    supabase=Depends(get_supabase_client),
):
    """Preview proposal as HTML page."""
    sim = _load_simulation(supabase, simulation_id)

    generator = ProposalGenerator()
    input_data = sim.get("input_data", {})
    result_data = sim.get("result", {})
    vehicle_info = _extract_vehicle_info(input_data)
    fund_info = {"fund_name": "カーチスファンド"}

    html = generator.generate_html_preview(result_data, vehicle_info, fund_info)
    return HTMLResponse(content=html)


# ---------------------------------------------------------------------------
# POST /export-design – Excel fund design document
# ---------------------------------------------------------------------------


@router.post("/export-design")
async def export_design_document(
    simulation_id: UUID,
    request: Request,
    user=Depends(require_permission("simulations", "read")),
    supabase=Depends(get_supabase_client),
):
    """Export fund design document as Excel.

    Sheets: サマリー, 車両一覧, プライシング根拠, キャッシュフロー,
    ステークホルダー.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl is not installed – Excel export unavailable",
        )

    sim = _load_simulation(supabase, simulation_id)

    input_data = sim.get("input_data", {})
    result_data = sim.get("result", {}) if isinstance(sim.get("result"), dict) else {}

    wb = openpyxl.Workbook()

    # -- shared styles -------------------------------------------------------
    header_fill = PatternFill(
        start_color="1A365D", end_color="1A365D", fill_type="solid"
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _style_header(ws, col_count: int) -> None:
        """Apply header styling to the first row."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # -- Sheet 1: サマリー (Summary) -----------------------------------------
    ws_summary = wb.active
    ws_summary.title = "サマリー"
    ws_summary.append(["項目", "値"])
    ws_summary.append(["メーカー", input_data.get("maker", "")])
    ws_summary.append(["車種", input_data.get("model", "")])
    ws_summary.append([
        "リース期間",
        f"{input_data.get('lease_term_months', '')} ヶ月",
    ])

    acq = result_data.get("acquisition", {})
    res = result_data.get("residual", {})
    lease = result_data.get("lease", {})

    ws_summary.append(["推奨買取価格", acq.get("recommended_acquisition_price", "")])
    ws_summary.append(["残価（Base）", res.get("scenario_base", "")])
    ws_summary.append(["月額リース料", lease.get("monthly_lease_fee", "")])
    ws_summary.append([
        "利益転換月",
        result_data.get("profit_conversion_month", ""),
    ])
    ws_summary.append(["評価", result_data.get("assessment", "")])
    _style_header(ws_summary, 2)
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 30

    # -- Sheet 2: 車両一覧 (Vehicle List) ------------------------------------
    ws_vehicles = wb.create_sheet("車両一覧")
    ws_vehicles.append([
        "メーカー",
        "車種",
        "登録年月",
        "走行距離(km)",
        "車両クラス",
        "ボディタイプ",
    ])
    ws_vehicles.append([
        input_data.get("maker", ""),
        input_data.get("model", ""),
        input_data.get("registration_year_month", ""),
        input_data.get("mileage_km", ""),
        input_data.get("vehicle_class", ""),
        input_data.get("body_type", ""),
    ])
    _style_header(ws_vehicles, 6)

    # -- Sheet 3: プライシング根拠 (Pricing Basis) ---------------------------
    ws_pricing = wb.create_sheet("プライシング根拠")
    ws_pricing.append(["パラメータ", "値"])
    ws_pricing.append(["市場中央値", acq.get("market_median_price", "")])
    ws_pricing.append(["サンプル数", acq.get("market_sample_count", "")])
    ws_pricing.append(["トレンド係数", acq.get("trend_factor", "")])
    ws_pricing.append(["安全マージン", acq.get("safety_margin_applied", "")])
    _style_header(ws_pricing, 2)
    ws_pricing.column_dimensions["A"].width = 22
    ws_pricing.column_dimensions["B"].width = 30

    # -- Sheet 4: キャッシュフロー (Cash Flow) --------------------------------
    ws_cf = wb.create_sheet("キャッシュフロー")
    ws_cf.append(["月", "帳簿価額", "累積収入", "累積費用", "累積損益", "NAV"])
    nav_curve = result_data.get("nav_curve", [])
    for point in nav_curve:
        ws_cf.append([
            point.get("month", ""),
            point.get("asset_book_value", ""),
            point.get("cumulative_lease_income", ""),
            point.get("cumulative_costs", ""),
            point.get("cumulative_profit", ""),
            point.get("nav", ""),
        ])
    _style_header(ws_cf, 6)

    # -- Sheet 5: ステークホルダー (Stakeholders) -----------------------------
    ws_sh = wb.create_sheet("ステークホルダー")
    ws_sh.append(["役割", "会社名", "代表者", "住所", "電話", "メール"])
    stakeholders = (
        supabase.table("deal_stakeholders")
        .select("*")
        .eq("simulation_id", str(simulation_id))
        .order("display_order")
        .execute()
    )
    for s in stakeholders.data:
        ws_sh.append([
            s.get("role_type", ""),
            s.get("company_name", ""),
            s.get("representative_name", ""),
            s.get("address", ""),
            s.get("phone", ""),
            s.get("email", ""),
        ])
    _style_header(ws_sh, 6)

    # -- Serialise workbook to bytes -----------------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"design_{input_data.get('maker', 'fund')}"
        f"_{input_data.get('model', '')}.xlsx"
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition(filename)},
    )
